"""
Flask Backend Server for CSV Regression Testing Web Interface
Integrates with the existing CSV regression testing tool
UPDATED TO SUPPORT TIMESTAMPED FILES
"""
from flask import Flask, request, jsonify, render_template_string, send_from_directory, send_file
from flask_cors import CORS
from werkzeug.utils import secure_filename

import os
import json
import tempfile
import shutil
from pathlib import Path
from datetime import datetime
import logging
import traceback
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import re

# Import your existing CSV regression testing classes
# Make sure to adjust the import path according to your project structure
from csv_regression_tester import (
    CSVRegressionTester,
    ValidationResult,
    ValidationIssue,
    ValidationReport
)

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

# Configuration
UPLOAD_FOLDER = 'uploads'
REFERENCE_FOLDER = 'reference_files'
SCHEMA_FILE = 'schema.json'
ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}

# Create necessary directories
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(REFERENCE_FOLDER, exist_ok=True)

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def extract_base_name_from_timestamped_file(filename: str) -> str:
    """
    Extract base name from timestamped filename
    Examples:
    - CashRecov_20250702184047.csv -> CashRecov
    - cashrecov_20250702184047.xlsx -> cashrecov
    - SomeFile_20250101120000.csv -> SomeFile
    """
    # Remove file extension first
    base_name = Path(filename).stem
    
    # Pattern to match timestamp suffixes like _20250702184047
    timestamp_pattern = r'_\d{14}$'  # Matches _YYYYMMDDHHMMSS at end
    
    # Remove timestamp suffix if present
    base_name = re.sub(timestamp_pattern, '', base_name)
    
    return base_name

def find_matching_reference_file(uploaded_filename: str) -> str:
    """
    Find matching reference file for uploaded file
    Handles both exact matches and timestamped files
    """
    # Extract base name from uploaded file
    base_name = extract_base_name_from_timestamped_file(uploaded_filename)
    
    # Try to find matching reference file
    supported_extensions = ['.csv', '.xlsx', '.xls']
    
    # Try exact match first
    for ext in supported_extensions:
        ref_file = f"{base_name}{ext}"
        if os.path.exists(os.path.join(REFERENCE_FOLDER, ref_file)):
            return ref_file
    
    # Try case-insensitive match
    for ext in supported_extensions:
        for filename in os.listdir(REFERENCE_FOLDER):
            if filename.lower().endswith(ext.lower()):
                file_base = Path(filename).stem
                if file_base.lower() == base_name.lower():
                    return filename
    
    return None

# def create_default_schema():
#     """Create a default schema file if it doesn't exist"""
#     default_schema = {
#         "default.csv": {
#             "columns": ["Order No", "Date", "Amount", "Customer", "Status"],
#             "types": {
#                 "Order No": "string",
#                 "Date": "date",
#                 "Amount": "float",
#                 "Customer": "string",
#                 "Status": "string"
#             },
#             "key_columns": ["Order No"],
#             "required_columns": ["Order No", "Date", "Amount"],
#             "row_count_policy": {
#                 "allow_more_rows": True,
#                 "allow_fewer_rows": False,
#                 "max_row_difference": 100
#             }
#         },
#         "sales_report.csv": {
#             "columns": ["Product", "Sales", "Quantity", "Date", "Region"],
#             "types": {
#                 "Product": "string",
#                 "Sales": "float",
#                 "Quantity": "int",
#                 "Date": "date",
#                 "Region": "string"
#             },
#             "key_columns": ["Product", "Date"],
#             "required_columns": ["Product", "Sales", "Date"],
#             "row_count_policy": {
#                 "allow_more_rows": True,
#                 "allow_fewer_rows": True,
#                 "max_row_difference": 50
#             }
#         }
#     }
    
#     if not os.path.exists(SCHEMA_FILE):
#         with open(SCHEMA_FILE, 'w') as f:
#             json.dump(default_schema, f, indent=2)
#         logger.info(f"Created default schema file: {SCHEMA_FILE}")

# def create_sample_reference_files():
#     """Create sample reference files for testing"""
#     # This is a placeholder - you should place your actual reference files here
#     sample_files = {
#         "sales_report.csv": """Product,Sales,Quantity,Date,Region
# Widget A,1000.50,100,2024-01-01,North
# Widget B,750.25,75,2024-01-01,South
# Widget C,1200.00,120,2024-01-02,East""",
        
#         "default.csv": """Order No,Date,Amount,Customer,Status
# ORD001,2024-01-01,299.99,Customer A,Completed
# ORD002,2024-01-02,499.50,Customer B,Pending
# ORD003,2024-01-03,150.00,Customer C,Completed"""
#     }
    
#     for filename, content in sample_files.items():
#         file_path = os.path.join(REFERENCE_FOLDER, filename)
#         if not os.path.exists(file_path):
#             with open(file_path, 'w') as f:
#                 f.write(content)
#             logger.info(f"Created sample reference file: {file_path}")

def create_schema_for_file(filename: str, reference_filename: str) -> dict:
    """
    Auto-detect schema from reference file
    """
    try:
        import pandas as pd
        ref_file_path = os.path.join(REFERENCE_FOLDER, reference_filename)
        
        # Try to read the reference file
        if ref_file_path.endswith('.csv'):
            ref_df = pd.read_csv(ref_file_path)
        elif ref_file_path.endswith(('.xlsx', '.xls')):
            ref_df = pd.read_excel(ref_file_path)
        else:
            raise ValueError(f"Unsupported file format: {reference_filename}")
        
        # Generate basic schema
        auto_schema = {
            "columns": ref_df.columns.tolist(),
            "types": {col: "string" for col in ref_df.columns},
            "key_columns": [ref_df.columns[0]] if len(ref_df.columns) > 0 else [],
            "required_columns": ref_df.columns.tolist()[:3],
            "row_count_policy": {
                "allow_more_rows": True,
                "allow_fewer_rows": True,
                "max_row_difference": 100
            }
        }
        
        # Try to detect better data types
        for col in ref_df.columns:
            if ref_df[col].dtype == 'int64':
                auto_schema["types"][col] = "int"
            elif ref_df[col].dtype == 'float64':
                auto_schema["types"][col] = "float"
            elif pd.api.types.is_datetime64_any_dtype(ref_df[col]):
                auto_schema["types"][col] = "date"
        
        logger.info(f"Auto-generated schema for {filename} using reference {reference_filename}")
        return auto_schema
        
    except Exception as e:
        logger.error(f"Could not auto-generate schema for {filename}: {e}")
        # Return minimal schema
        return {
            "columns": [],
            "types": {},
            "key_columns": [],
            "required_columns": [],
            "row_count_policy": {
                "allow_more_rows": True,
                "allow_fewer_rows": True,
                "max_row_difference": 100
            }
        }

def validation_report_to_dict(report: ValidationReport) -> dict:
    """Convert ValidationReport object to dictionary for JSON serialization"""
    return {
        'file_name': report.file_name,
        'reference_file': report.reference_file,
        'generated_file': report.generated_file,
        'timestamp': report.timestamp.isoformat(),
        'overall_result': report.overall_result.value,
        'issues': [
            {
                'issue_type': issue.issue_type,
                'severity': issue.severity.value,
                'message': issue.message,
                'row_index': issue.row_index,
                'column': issue.column,
                'expected_value': str(issue.expected_value) if issue.expected_value is not None else None,
                'actual_value': str(issue.actual_value) if issue.actual_value is not None else None
            }
            for issue in report.issues
        ],
        'summary': report.summary
    }

@app.route('/')
def index():
    """Serve the main web interface"""
    # You can serve the HTML file directly or render it as a template
    try:
        with open('web_interface.html', 'r',encoding='utf-8') as f:
            return f.read()
    except FileNotFoundError:
        return '''
        <!DOCTYPE html>
        <html>
        <head>
            <title>CSV Regression Testing Tool</title>
            <style>
                body { font-family: Arial, sans-serif; margin: 40px; text-align: center; }
                .error { color: red; padding: 20px; border: 1px solid red; border-radius: 5px; }
            </style>
        </head>
        <body>
            <h1>CSV Regression Testing Tool</h1>
            <div class="error">
                <p>Web interface file not found. Please ensure 'web_interface.html' is in the same directory as this server.</p>
                <p>You can still use the API endpoints:</p>
                <ul>
                    <li>POST /api/upload - Upload and test files</li>
                    <li>GET /api/schema - Get current schema</li>
                    <li>POST /api/schema - Update schema</li>
                    <li>GET /api/reference-files - List reference files</li>
                    <li>GET /api/status - Server status</li>
                </ul>
            </div>
        </body>
        </html>
        '''

@app.route('/api/upload', methods=['POST'])
def upload_and_test():
    """Handle file upload and run regression tests"""
    try:
        # Check if files were uploaded
        if 'files' not in request.files:
            return jsonify({'success': False, 'error': 'No files uploaded'}), 400
        
        files = request.files.getlist('files')
        if not files or all(f.filename == '' for f in files):
            return jsonify({'success': False, 'error': 'No files selected'}), 400
        
        # Create temporary directory for uploaded files
        temp_dir = tempfile.mkdtemp()
        logger.info(f"Created temporary directory: {temp_dir}")
        
        try:
            # Save uploaded files
            uploaded_files = []
            for file in files:
                if file and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    file_path = os.path.join(temp_dir, filename)
                    file.save(file_path)
                    uploaded_files.append(filename)
                    logger.info(f"Saved uploaded file: {filename}")
            
            if not uploaded_files:
                return jsonify({'success': False, 'error': 'No valid files uploaded'}), 400
            
            # Load existing schema
            schema_config = {}
            if os.path.exists(SCHEMA_FILE):
                with open(SCHEMA_FILE, 'r') as f:
                    schema_config = json.load(f)
            
            # Initialize CSV regression tester
            tester = CSVRegressionTester(
                schema_file=SCHEMA_FILE,
                reference_dir=REFERENCE_FOLDER,
                generated_dir=temp_dir
            )
            
            # Run tests for each uploaded file
            reports = []
            for filename in uploaded_files:
                try:
                    logger.info(f"Processing file: {filename}")
                    
                    # Find matching reference file
                    reference_filename = find_matching_reference_file(filename)
                    
                    if not reference_filename:
                        # No matching reference file found
                        base_name = extract_base_name_from_timestamped_file(filename)
                        error_report = ValidationReport(
                            file_name=filename,
                            reference_file=f"NOT FOUND: {base_name}.*",
                            generated_file=f"uploads/{filename}",
                            overall_result=ValidationResult.FAIL
                        )
                        error_report.add_issue(ValidationIssue(
                            issue_type="reference_file_not_found",
                            severity=ValidationResult.FAIL,
                            message=f"No reference file found for base name: {base_name}"
                        ))
                        error_report.generate_summary()
                        reports.append(validation_report_to_dict(error_report))
                        continue
                    
                    logger.info(f"Found matching reference file: {reference_filename}")
                    
                    # Check if we have a schema for this file
                    # Try the uploaded filename first, then the reference filename
                    schema_key = None
                    if filename in schema_config:
                        schema_key = filename
                    elif reference_filename in schema_config:
                        schema_key = reference_filename
                    else:
                        # Try base name variations
                        base_name = extract_base_name_from_timestamped_file(filename)
                        for key in schema_config.keys():
                            if extract_base_name_from_timestamped_file(key).lower() == base_name.lower():
                                schema_key = key
                                break
                    
                    if not schema_key:
                        # Auto-generate schema from reference file
                        auto_schema = create_schema_for_file(filename, reference_filename)
                        schema_config[filename] = auto_schema
                        tester.schema_config[filename] = auto_schema
                        
                        # Update schema file
                        with open(SCHEMA_FILE, 'w') as f:
                            json.dump(schema_config, f, indent=2)
                        
                        logger.info(f"Created auto-schema for {filename}")
                    else:
                        # Use existing schema but ensure it's loaded into tester
                        tester.schema_config[filename] = schema_config[schema_key]
                        logger.info(f"Using existing schema for {filename} (key: {schema_key})")
                    
                    # Run test for this file
                    report = tester.test_file(filename)
                    reports.append(validation_report_to_dict(report))
                    logger.info(f"Completed test for {filename}: {report.overall_result.value}")
                
                except Exception as e:
                    logger.error(f"Error testing file {filename}: {e}")
                    logger.error(traceback.format_exc())
                    error_report = ValidationReport(
                        file_name=filename,
                        reference_file=f"reference_files/{filename}",
                        generated_file=f"uploads/{filename}",
                        overall_result=ValidationResult.FAIL
                    )
                    error_report.add_issue(ValidationIssue(
                        issue_type="processing_error",
                        severity=ValidationResult.FAIL,
                        message=f"Error processing file: {str(e)}"
                    ))
                    error_report.generate_summary()
                    reports.append(validation_report_to_dict(error_report))
            
            if not reports:
                return jsonify({
                    'success': False, 
                    'error': 'No files could be processed. Check if reference files exist and schemas are configured.'
                }), 400
            
            return jsonify({
                'success': True,
                'data': reports,
                'message': f'Processed {len(reports)} files successfully'
            })
        
        except Exception as e:
            logger.error(f"Error during file processing: {e}")
            logger.error(traceback.format_exc())
            return jsonify({
                'success': False,
                'error': f'Error processing files: {str(e)}'
            }), 500
        
        finally:
            # Clean up temporary directory
            try:
                shutil.rmtree(temp_dir)
                logger.info(f"Cleaned up temporary directory: {temp_dir}")
            except Exception as e:
                logger.error(f"Error cleaning up temp directory: {e}")
    
    except Exception as e:
        logger.error(f"Unexpected error in upload_and_test: {e}")
        logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': f'Unexpected error: {str(e)}'
        }), 500

@app.route('/api/schema', methods=['GET'])
def get_schema():
    """Get current schema configuration"""
    try:
        if os.path.exists(SCHEMA_FILE):
            with open(SCHEMA_FILE, 'r') as f:
                schema = json.load(f)
            return jsonify({'success': True, 'schema': schema})
        else:
            return jsonify({'success': False, 'error': 'Schema file not found'}), 404
    except Exception as e:
        logger.error(f"Error reading schema: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/schema', methods=['POST'])
def update_schema():
    """Update schema configuration"""
    try:
        schema_data = request.get_json()
        if not schema_data:
            return jsonify({'success': False, 'error': 'No schema data provided'}), 400
        
        # Validate schema structure (basic validation)
        for file_name, config in schema_data.items():
            required_keys = ['columns', 'types', 'key_columns', 'required_columns']
            for key in required_keys:
                if key not in config:
                    return jsonify({
                        'success': False, 
                        'error': f'Missing required key "{key}" in schema for {file_name}'
                    }), 400
        
        # Save updated schema
        with open(SCHEMA_FILE, 'w') as f:
            json.dump(schema_data, f, indent=2)
        
        logger.info("Schema updated successfully")
        return jsonify({'success': True, 'message': 'Schema updated successfully'})
    
    except Exception as e:
        logger.error(f"Error updating schema: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/reference-files', methods=['GET'])
def list_reference_files():
    """List available reference files"""
    try:
        reference_files = []
        if os.path.exists(REFERENCE_FOLDER):
            for filename in os.listdir(REFERENCE_FOLDER):
                if allowed_file(filename):
                    file_path = os.path.join(REFERENCE_FOLDER, filename)
                    file_info = {
                        'name': filename,
                        'size': os.path.getsize(file_path),
                        'modified': datetime.fromtimestamp(os.path.getmtime(file_path)).isoformat()
                    }
                    reference_files.append(file_info)
        
        return jsonify({'success': True, 'files': reference_files})
    
    except Exception as e:
        logger.error(f"Error listing reference files: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/reference-files/<filename>', methods=['GET'])
def download_reference_file(filename):
    """Download a specific reference file"""
    try:
        filename = secure_filename(filename)
        return send_from_directory(REFERENCE_FOLDER, filename, as_attachment=True)
    except Exception as e:
        logger.error(f"Error downloading reference file: {e}")
        return jsonify({'success': False, 'error': str(e)}), 404

@app.route('/api/reference-files/<filename>', methods=['POST'])
def upload_reference_file(filename):
    """Upload/update a reference file"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'success': False, 'error': 'Invalid file type'}), 400
        
        filename = secure_filename(filename)
        file_path = os.path.join(REFERENCE_FOLDER, filename)
        file.save(file_path)
        
        logger.info(f"Reference file uploaded: {filename}")
        return jsonify({'success': True, 'message': f'Reference file {filename} uploaded successfully'})
    
    except Exception as e:
        logger.error(f"Error uploading reference file: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/status', methods=['GET'])
def server_status():
    """Get server status and configuration"""
    try:
        status = {
            'server_time': datetime.now().isoformat(),
            'schema_file_exists': os.path.exists(SCHEMA_FILE),
            'reference_folder_exists': os.path.exists(REFERENCE_FOLDER),
            'upload_folder_exists': os.path.exists(UPLOAD_FOLDER),
            'reference_files_count': len([f for f in os.listdir(REFERENCE_FOLDER) 
                                        if allowed_file(f)]) if os.path.exists(REFERENCE_FOLDER) else 0,
            'allowed_extensions': list(ALLOWED_EXTENSIONS)
        }
        
        return jsonify({'success': True, 'status': status})
    
    except Exception as e:
        logger.error(f"Error getting server status: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/test-single', methods=['POST'])
def test_single_file():
    """Test a single file without upload (for existing files)"""
    try:
        data = request.get_json()
        if not data or 'filename' not in data:
            return jsonify({'success': False, 'error': 'Filename required'}), 400
        
        filename = data['filename']
        
        # Check if file exists in reference folder
        ref_file = os.path.join(REFERENCE_FOLDER, filename)
        if not os.path.exists(ref_file):
            return jsonify({'success': False, 'error': f'Reference file {filename} not found'}), 404
        
        # Initialize tester
        tester = CSVRegressionTester(
            schema_file=SCHEMA_FILE,
            reference_dir=REFERENCE_FOLDER,
            generated_dir=REFERENCE_FOLDER  # Use reference as generated for demo
        )
        
        # Run test
        report = tester.test_file(filename)
        return jsonify({
            'success': True,
            'data': validation_report_to_dict(report)
        })
    
    except Exception as e:
        logger.error(f"Error testing single file: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.errorhandler(404)
def not_found(error):
    return jsonify({'success': False, 'error': 'Endpoint not found'}), 404

@app.errorhandler(500)
def internal_error(error):
    logger.error(f"Internal server error: {error}")
    return jsonify({'success': False, 'error': 'Internal server error'}), 500

# Initialize on startup
def initialize_server():
    """Initialize server with default configuration"""
    try:
        # create_default_schema()
        # create_sample_reference_files()
        logger.info("Server initialized successfully")
    except Exception as e:
        logger.error(f"Error initializing server: {e}")





@app.route('/api/validate-extension', methods=['POST'])
def validate_extension():
    try:
        data = request.json
        uploaded_filename = data.get('baseName', '')
        uploaded_extension = data.get('uploadedExtension', '')
        
        logger.info(f"Validating extension for uploaded_filename: {uploaded_filename}, extension: {uploaded_extension}")
        
        # Use the same configuration as other endpoints in your Flask app
        tester = CSVRegressionTester(
            schema_file=SCHEMA_FILE,           # Use Flask app constants
            reference_dir=REFERENCE_FOLDER,    # Use Flask app constants  
            generated_dir=REFERENCE_FOLDER     # Use Flask app constants (same as in test_single_file endpoint)
        )
        
        # Extract the base name from the uploaded filename using the CSV regression function
        base_name = tester._extract_base_name_from_timestamped_file(uploaded_filename)
        logger.info(f"Extracted base name: '{base_name}' from uploaded filename: '{uploaded_filename}'")
        
        # Debug: List reference files and their extracted base names
        if os.path.exists(REFERENCE_FOLDER):
            ref_files = [f for f in os.listdir(REFERENCE_FOLDER) if allowed_file(f)]
            logger.info(f"Reference files and their base names:")
            for ref_file in ref_files:
                ref_base = tester._extract_base_name_from_timestamped_file(ref_file)
                logger.info(f"  '{ref_file}' -> '{ref_base}'")
        
        # Find reference file with extension info using the CSV regression function
        ref_file, extension_warning, other_extension_files = tester._find_reference_file_with_extension_info(
            base_name, uploaded_extension
        )
        
        if not ref_file:
            # Check if there are any files with same base name but different extensions
            all_potential_files = []
            for ext in ['.csv', '.xlsx', '.xls']:
                for file_path in tester.reference_dir.glob(f"*{ext}"):
                    file_base = tester._extract_base_name_from_timestamped_file(file_path.name)
                    if file_base.lower() == base_name.lower():
                        all_potential_files.append(file_path)
            
            logger.info(f"Looking for base name '{base_name}', found potential files: {[str(f) for f in all_potential_files]}")
            
            if all_potential_files:
                extensions_found = [f.suffix for f in all_potential_files]
                error_message = f"Please upload a file with extension {extensions_found[0]} to match the reference file for '{base_name}'. Available reference extensions: {extensions_found}"
                return jsonify({
                    'success': False,
                    'error': error_message,
                    'available_extensions': extensions_found
                })
            else:
                return jsonify({
                    'success': False,
                    'error': f"No reference file found for '{base_name}' with any supported extension (.csv, .xlsx, .xls)"
                })
        
        # File exists with matching extension
        return jsonify({
            'success': True,
            'message': 'Extension validation passed',
            'reference_file': str(ref_file),
            'extension_warning': extension_warning,
            'extracted_base_name': base_name  # Include for debugging
        })
        
    except Exception as e:
        logger.error(f"Extension validation error: {e}")
        return jsonify({
            'success': False,
            'error': f'Extension validation error: {str(e)}'
        })



# reports  code  starts here

@app.route('/reports.html')
def reports_page():
    """Serve the reports HTML page"""
    return send_from_directory('.', 'reports.html')
def get_file_size(filepath):
    """Get file size in bytes"""
    return os.path.getsize(filepath)

def format_file_size(size_bytes):
    """Format file size in human readable format"""
    if size_bytes == 0:
        return "0B"
    size_names = ["B", "KB", "MB", "GB"]
    i = 0
    while size_bytes >= 1024 and i < len(size_names) - 1:
        size_bytes /= 1024.0
        i += 1
    return f"{size_bytes:.1f}{size_names[i]}"

@app.route('/api/reference-reports', methods=['GET'])
def get_reference_reports():
    """Get list of all reference reports"""
    try:
        reports = []
        
        if not os.path.exists(REFERENCE_FOLDER):
            return jsonify({'success': True, 'reports': []})
        
        for filename in os.listdir(REFERENCE_FOLDER):
            if allowed_file(filename):
                filepath = os.path.join(REFERENCE_FOLDER, filename)
                file_stats = os.stat(filepath)
                
                report_info = {
                    'name': filename,
                    'extension': filename.rsplit('.', 1)[1].lower(),
                    'size': file_stats.st_size,
                    'modified': datetime.fromtimestamp(file_stats.st_mtime).isoformat(),
                    'description': None  # You can add descriptions later if needed
                }
                reports.append(report_info)
        
        # Sort by modification date (newest first)
        reports.sort(key=lambda x: x['modified'], reverse=True)
        
        logger.info(f"Retrieved {len(reports)} reference reports")
        return jsonify({'success': True, 'reports': reports})
    
    except Exception as e:
        logger.error(f"Error retrieving reference reports: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/preview-report/<filename>', methods=['GET'])
def preview_report(filename):
    """Preview Excel report data"""
    try:
        secure_name = secure_filename(filename)
        filepath = os.path.join(REFERENCE_FOLDER, secure_name)
        
        if not os.path.exists(filepath):
            logger.warning(f"File not found for preview: {secure_name}")
            return jsonify({'success': False, 'error': 'File not found'}), 404
        
        file_extension = secure_name.rsplit('.', 1)[1].lower()
        
        if file_extension not in ['xlsx', 'xls']:
            return jsonify({'success': False, 'error': 'Preview only available for Excel files'}), 400
        
        # Load workbook with better error handling
        try:
            workbook = load_workbook(filepath, data_only=True)
        except Exception as e:
            logger.error(f"Error loading workbook {secure_name}: {str(e)}")
            return jsonify({'success': False, 'error': f'Unable to read Excel file: {str(e)}'}), 500
        
        sheets_data = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Get sheet dimensions
            max_row = sheet.max_row or 0
            max_col = sheet.max_column or 0
            
            # Handle empty sheets
            if max_row == 0 or max_col == 0:
                sheet_info = {
                    'name': sheet_name,
                    'rows': 0,
                    'columns': 0,
                    'headers': [],
                    'data': []
                }
                sheets_data.append(sheet_info)
                continue
            
            # Get headers (first row)
            headers = []
            for col in range(1, max_col + 1):
                cell_value = sheet.cell(row=1, column=col).value
                headers.append(str(cell_value) if cell_value is not None else f"Column {col}")
            
            # Get data (skip header row)
            data = []
            for row in range(2, min(max_row + 1, 102)):  # Limit to 100 data rows
                row_data = []
                for col in range(1, max_col + 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value is None:
                        row_data.append("")
                    elif isinstance(cell_value, (int, float)):
                        row_data.append(str(cell_value))
                    elif isinstance(cell_value, datetime):
                        row_data.append(cell_value.strftime('%Y-%m-%d %H:%M:%S'))
                    else:
                        row_data.append(str(cell_value))
                data.append(row_data)
            
            sheet_info = {
                'name': sheet_name,
                'rows': max_row - 1 if max_row > 1 else 0,  # Excluding header
                'columns': max_col,
                'headers': headers,
                'data': data
            }
            sheets_data.append(sheet_info)
        
        logger.info(f"Generated preview for {secure_name} with {len(sheets_data)} sheets")
        return jsonify({
            'success': True,
            'data': {
                'sheets': sheets_data
            }
        })
    
    except Exception as e:
        logger.error(f"Error previewing report {filename}: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/download-report/<filename>', methods=['GET'])
def download_report(filename):
    """Download a reference report"""
    try:
        secure_name = secure_filename(filename)
        filepath = os.path.join(REFERENCE_FOLDER, secure_name)
        
        if not os.path.exists(filepath):
            logger.warning(f"File not found for download: {secure_name}")
            return jsonify({'success': False, 'error': 'File not found'}), 404
        
        logger.info(f"Downloading report: {secure_name}")
        return send_file(
            filepath,
            as_attachment=True,
            download_name=secure_name,
            mimetype='application/octet-stream'
        )
    
    except Exception as e:
        logger.error(f"Error downloading report {filename}: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/analyze-report/<filename>', methods=['GET'])
def analyze_report(filename):
    """Analyze Excel report and return detailed information"""
    try:
        secure_name = secure_filename(filename)
        filepath = os.path.join(REFERENCE_FOLDER, secure_name)
        
        if not os.path.exists(filepath):
            logger.warning(f"File not found for analysis: {secure_name}")
            return jsonify({'success': False, 'error': 'File not found'}), 404
        
        file_extension = secure_name.rsplit('.', 1)[1].lower()
        
        if file_extension not in ['xlsx', 'xls']:
            return jsonify({'success': False, 'error': 'Analysis only available for Excel files'}), 400
        
        # Load workbook
        workbook = load_workbook(filepath, data_only=True)
        file_size = get_file_size(filepath)
        
        sheets_analysis = []
        total_rows = 0
        total_columns = 0
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Get sheet dimensions
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            # Get column headers
            column_headers = []
            for col in range(1, max_col + 1):
                cell_value = sheet.cell(row=1, column=col).value
                column_headers.append(str(cell_value) if cell_value is not None else f"Column {col}")
            
            # Analyze data types (sample first few rows)
            data_types = set()
            for row in range(2, min(max_row + 1, 12)):  # Sample 10 rows
                for col in range(1, max_col + 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value is not None:
                        if isinstance(cell_value, (int, float)):
                            data_types.add('Number')
                        elif isinstance(cell_value, datetime):
                            data_types.add('Date')
                        else:
                            data_types.add('Text')
            
            # Fixed: Use different key names to avoid duplication
            sheet_analysis = {
                'name': sheet_name,
                'rows': max_row - 1 if max_row > 1 else 0,  # Excluding header
                'columns': max_col,  # Number of columns
                'dataTypes': list(data_types) if data_types else ['Empty'],
                'columnHeaders': column_headers  # Array of column names
            }
            sheets_analysis.append(sheet_analysis)
            
            total_rows += max_row - 1 if max_row > 1 else 0
            total_columns += max_col
        
        analysis = {
            'fileSize': file_size,
            'sheetsCount': len(workbook.sheetnames),
            'totalRows': total_rows,
            'totalColumns': total_columns,
            'sheets': sheets_analysis
        }
        
        logger.info(f"Analyzed report {secure_name}: {len(sheets_analysis)} sheets, {total_rows} total rows")
        return jsonify({
            'success': True,
            'analysis': analysis
        })
    
    except Exception as e:
        logger.error(f"Error analyzing report {filename}: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/upload-reference-report', methods=['POST'])
def upload_reference_report():
    """Upload a new reference report"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(REFERENCE_FOLDER, filename)
            
            # Check if file already exists
            if os.path.exists(filepath):
                return jsonify({'success': False, 'error': 'File already exists'}), 409
            
            file.save(filepath)
            
            # Get file info
            file_stats = os.stat(filepath)
            report_info = {
                'name': filename,
                'extension': filename.rsplit('.', 1)[1].lower(),
                'size': file_stats.st_size,
                'modified': datetime.fromtimestamp(file_stats.st_mtime).isoformat()
            }
            
            logger.info(f"Uploaded reference report: {filename}")
            return jsonify({
                'success': True,
                'message': f'Report "{filename}" uploaded successfully',
                'report': report_info
            })
        
        return jsonify({'success': False, 'error': 'Invalid file type'}), 400
    
    except Exception as e:
        logger.error(f"Error uploading reference report: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/delete-reference-report/<filename>', methods=['DELETE'])
def delete_reference_report(filename):
    """Delete a reference report"""
    try:
        secure_name = secure_filename(filename)
        filepath = os.path.join(REFERENCE_FOLDER, secure_name)
        
        if not os.path.exists(filepath):
            logger.warning(f"File not found for deletion: {secure_name}")
            return jsonify({'success': False, 'error': 'File not found'}), 404
        
        os.remove(filepath)
        
        logger.info(f"Deleted reference report: {secure_name}")
        return jsonify({
            'success': True,
            'message': f'Report "{secure_name}" deleted successfully'
        })
    
    except Exception as e:
        logger.error(f"Error deleting reference report {filename}: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/preview-csv/<filename>', methods=['GET'])
def preview_csv(filename):
    """Preview CSV file data"""
    try:
        secure_name = secure_filename(filename)
        filepath = os.path.join(REFERENCE_FOLDER, secure_name)
        
        if not os.path.exists(filepath):
            logger.warning(f"CSV file not found for preview: {secure_name}")
            return jsonify({'success': False, 'error': 'File not found'}), 404
        
        # Read CSV file with better error handling
        try:
            df = pd.read_csv(filepath, encoding='utf-8')
        except UnicodeDecodeError:
            # Try with different encoding
            try:
                df = pd.read_csv(filepath, encoding='latin1')
            except:
                df = pd.read_csv(filepath, encoding='cp1252')
        
        # Handle empty DataFrames
        if df.empty:
            return jsonify({
                'success': True,
                'data': {
                    'headers': [],
                    'data': [],
                    'totalRows': 0,
                    'totalColumns': 0
                }
            })
        
        # Limit to first 100 rows for preview
        preview_df = df.head(100)
        
        # Convert to JSON-serializable format
        data_list = []
        for _, row in preview_df.iterrows():
            row_data = []
            for value in row:
                if pd.isna(value):
                    row_data.append("")
                else:
                    row_data.append(str(value))
            data_list.append(row_data)
        
        logger.info(f"Generated CSV preview for {secure_name}: {len(df)} total rows")
        return jsonify({
            'success': True,
            'data': {
                'headers': df.columns.tolist(),
                'data': data_list,
                'totalRows': len(df),
                'totalColumns': len(df.columns)
            }
        })
    
    except Exception as e:
        logger.error(f"Error previewing CSV {filename}: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/find-matching-reference/<filename>', methods=['GET'])
def find_matching_reference(filename):
    """Find matching reference file for uploaded file using existing logic"""
    try:
        secure_name = secure_filename(filename)
        matching_file = find_matching_reference_file(secure_name)
        
        if matching_file:
            logger.info(f"Found matching reference file for {secure_name}: {matching_file}")
            return jsonify({
                'success': True,
                'matchingFile': matching_file,
                'message': f'Found matching reference file: {matching_file}'
            })
        else:
            logger.info(f"No matching reference file found for {secure_name}")
            return jsonify({
                'success': False,
                'message': f'No matching reference file found for {secure_name}'
            })
    
    except Exception as e:
        logger.error(f"Error finding matching reference for {filename}: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/reference-reports/stats', methods=['GET'])
def get_reference_reports_stats():
    """Get statistics about reference reports"""
    try:
        if not os.path.exists(REFERENCE_FOLDER):
            return jsonify({'success': True, 'stats': {'total': 0, 'byType': {}}})
        
        stats = {
            'total': 0,
            'byType': {'csv': 0, 'xlsx': 0, 'xls': 0},
            'totalSize': 0
        }
        
        for filename in os.listdir(REFERENCE_FOLDER):
            if allowed_file(filename):
                filepath = os.path.join(REFERENCE_FOLDER, filename)
                file_extension = filename.rsplit('.', 1)[1].lower()
                
                stats['total'] += 1
                stats['byType'][file_extension] += 1
                stats['totalSize'] += get_file_size(filepath)
        
        logger.info(f"Reference reports stats: {stats['total']} total files")
        return jsonify({'success': True, 'stats': stats})
    
    except Exception as e:
        logger.error(f"Error getting reference reports stats: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/file-info/<filename>', methods=['GET'])
def get_file_info(filename):
    """Get detailed information about a file"""
    try:
        secure_name = secure_filename(filename)
        filepath = os.path.join(REFERENCE_FOLDER, secure_name)
        
        if not os.path.exists(filepath):
            return jsonify({'success': False, 'error': 'File not found'}), 404
        
        file_stats = os.stat(filepath)
        file_extension = secure_name.rsplit('.', 1)[1].lower()
        
        file_info = {
            'name': secure_name,
            'extension': file_extension,
            'size': file_stats.st_size,
            'sizeFormatted': format_file_size(file_stats.st_size),
            'modified': datetime.fromtimestamp(file_stats.st_mtime).isoformat(),
            'modifiedFormatted': datetime.fromtimestamp(file_stats.st_mtime).strftime('%Y-%m-%d %H:%M:%S'),
            'canPreview': file_extension in ['xlsx', 'xls', 'csv'],
            'canAnalyze': file_extension in ['xlsx', 'xls']
        }
        
        return jsonify({
            'success': True,
            'fileInfo': file_info
        })
    
    except Exception as e:
        logger.error(f"Error getting file info for {filename}: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

if __name__ == '__main__':
    # Initialize server
    initialize_server()
    
    # Run the Flask app
    app.run(
        host='0.0.0.0',
        port=5000,
        debug=True,
        threaded=True
    )
